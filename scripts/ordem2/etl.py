"""
Roots of Brazil — ETL da Ordem 2.

Lê o Corpus_Fundador_v1.1.xlsx (READ-ONLY, nunca escrito) e carrega os dados
reais em um banco SQLite local, seguindo o DDL de schemas/ddl_sqlite.sql.

Gera uuid v4 e slug no momento da carga (não existem na fonte v1.1).
created_at/updated_at = timestamp da Ata de Homologação v1.1 (2026-08-05), version = 1.
Campos ainda não populados (Familia/Ordem/Grupo/Macrogrupo, geodados, peso, LIV_ID)
recebem os valores sentinela documentados no Dicionário v1.2 — nunca inventados.

Se qualquer contagem pós-carga divergir do Relatório de Auditoria Sprint 2
(381 objetos, 1.585 relações, 0 duplicidades, 18 órfãos), o script PARA e
reporta a divergência — não ajusta dado para bater número.
"""

from __future__ import annotations

import re
import sqlite3
import sys
import unicodedata
import uuid
from pathlib import Path

import openpyxl

CORPUS_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Corpus_Fundador_v1.1.xlsx")
DB_PATH = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("roots_of_brazil_dev.db")
DDL_PATH = Path(__file__).resolve().parents[2] / "schemas" / "ddl_sqlite.sql"

TIMESTAMP_HOMOLOGACAO = "2026-08-05T00:00:00Z"
SENTINELA_TAXONOMIA = "não classificado"


def slugify(nome: str, seen: set[str]) -> str:
    nfkd = unicodedata.normalize("NFKD", nome)
    ascii_ = nfkd.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_.lower()).strip("-")
    base = slug
    n = 2
    while slug in seen:
        slug = f"{base}-{n}"
        n += 1
    seen.add(slug)
    return slug


def sheet_rows(ws, header_row: int = 1):
    headers = [c.value for c in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(headers, row))


def main() -> None:
    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(DB_PATH)
    conn.executescript(DDL_PATH.read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(CORPUS_PATH, data_only=True)  # read-only, arquivo original nunca escrito
    slugs_por_tabela: dict[str, set[str]] = {}

    def novo_slug(tabela: str, nome: str) -> str:
        seen = slugs_por_tabela.setdefault(tabela, set())
        return slugify(nome, seen)

    # --- Ingredientes ---
    n_ing = 0
    for r in sheet_rows(wb["1. Catálogo Ingredientes"]):
        slug = novo_slug("ingrediente", r["Nome principal"])
        conn.execute(
            """INSERT INTO ingrediente (id, uuid, slug, created_at, updated_at, version,
                nome_principal, nomes_regionais, categoria, subcategoria, classe,
                familia, ordem_taxonomica, grupo, macrogrupo, origem_texto, estado_regiao, bioma_texto,
                confiabilidade, n_livros_fonte, n_citacoes, origem_registro, liv_id,
                nome_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Nome principal"], r["Nomes regionais"], r["Categoria"], r["Subcategoria"], r["Classe"],
             SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA,
             r["Origem"], r["Estado/Região"], r["Bioma"],
             r["Confiabilidade"], r["Nº de Livros-fonte"], r["Nº de Citações"], r["Origem do registro"], None,
             r["Nome principal"]),
        )
        n_ing += 1

    # --- Receitas ---
    n_rec = 0
    for r in sheet_rows(wb["2. Catálogo Receitas"]):
        slug = novo_slug("receita", r["Nome"])
        conn.execute(
            """INSERT INTO receita (id, uuid, slug, created_at, updated_at, version,
                nome, categoria, subcategoria, classe, familia, ordem_taxonomica, grupo, macrogrupo,
                estado, regiao, influencia_cultural, n_versoes_catalogadas, livros_fonte, liv_id,
                confiabilidade, nome_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Nome"], r["Categoria"], r["Subcategoria"], r["Classe"],
             SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA,
             r["Estado"], r["Região"], r["Influência cultural"], r["Nº de versões catalogadas"],
             r["Livros-fonte"], None, r["Confiabilidade"], r["Nome"]),
        )
        n_rec += 1

    # --- Técnicas ---
    n_tec = 0
    for r in sheet_rows(wb["3. Catálogo Técnicas"]):
        slug = novo_slug("tecnica", r["Nome"])
        conn.execute(
            """INSERT INTO tecnica (id, uuid, slug, created_at, updated_at, version,
                nome, descricao, ingredientes_utilizados, receitas, origem_cultural, livros_fonte, liv_id,
                confiabilidade, categoria, subcategoria, classe, familia, ordem_taxonomica, grupo, macrogrupo,
                nome_pt, descricao_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Nome"], r["Descrição"], r["Ingredientes utilizados"], r["Receitas"], r["Origem cultural"],
             r["Livros-fonte"], None, r["Confiabilidade"], r["Categoria"], r["Subcategoria"], r["Classe"],
             SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA,
             r["Nome"], r["Descrição"]),
        )
        n_tec += 1

    # --- Povos ---
    n_pov = 0
    for r in sheet_rows(wb["4. Catálogo Povos"]):
        slug = novo_slug("povo", r["Povo"])
        conn.execute(
            """INSERT INTO povo (id, uuid, slug, created_at, updated_at, version,
                povo, regiao, ingredientes_associados, receitas, praticas_culinarias, livros_fonte, liv_id,
                confiabilidade, categoria, subcategoria, classe, familia, ordem_taxonomica, grupo, macrogrupo,
                nome_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Povo"], r["Região"], r["Ingredientes associados"], r["Receitas"], r["Práticas culinárias"],
             r["Livros-fonte"], None, r["Confiabilidade"], r["Categoria"], r["Subcategoria"], r["Classe"],
             SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA,
             r["Povo"]),
        )
        n_pov += 1

    # --- Territórios ---
    n_ter = 0
    for r in sheet_rows(wb["5. Catálogo Territórios"]):
        slug = novo_slug("territorio", r["Estado"])
        conn.execute(
            """INSERT INTO territorio (id, uuid, slug, created_at, updated_at, version,
                estado, bioma_texto, ingredientes, receitas, produtos_tradicionais, livros_fonte, liv_id,
                confiabilidade, categoria, subcategoria, classe, familia, ordem_taxonomica, grupo, macrogrupo,
                nome_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Estado"], r["Bioma"], r["Ingredientes"], r["Receitas"], r["Produtos tradicionais"],
             r["Livros-fonte"], None, r["Confiabilidade"], r["Categoria"], r["Subcategoria"], r["Classe"],
             SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA,
             r["Estado"]),
        )
        n_ter += 1

    # --- Patrimônio ---
    n_pat = 0
    for r in sheet_rows(wb["6. Catálogo Patrimônio"]):
        slug = novo_slug("patrimonio", r["Elemento"])
        conn.execute(
            """INSERT INTO patrimonio (id, uuid, slug, created_at, updated_at, version,
                categoria, elemento, descricao, povo_regiao_relacionada, livros_fonte, liv_id,
                confiabilidade, subcategoria, classe, familia, ordem_taxonomica, grupo, macrogrupo,
                nome_pt, descricao_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Categoria"], r["Elemento"], r["Descrição"], r["Povo/Região relacionada"],
             r["Livros-fonte"], None, r["Confiabilidade"], r["Subcategoria"], r["Classe"],
             SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA, SENTINELA_TAXONOMIA,
             r["Elemento"], r["Descrição"]),
        )
        n_pat += 1

    # --- Biomas (cabeçalho na linha 3) ---
    n_bio = 0
    for r in sheet_rows(wb["12. Catálogo Biomas"], header_row=3):
        slug = novo_slug("bioma", r["Nome"])
        oficial_ibge = 0 if "extra-oficial" in (r["Nome"] or "").lower() or "Litoral" in (r["Nome"] or "") else 1
        conn.execute(
            """INSERT INTO bioma (id, uuid, slug, created_at, updated_at, version,
                nome, descricao, territorios_associados_n, ingredientes_associados_n, fonte, oficial_ibge,
                nome_pt, descricao_pt) VALUES (?,?,?,?,?,1,?,?,?,?,?,?,?,?)""",
            (r["ID"], str(uuid.uuid4()), slug, TIMESTAMP_HOMOLOGACAO, TIMESTAMP_HOMOLOGACAO,
             r["Nome"], r["Descrição"], r["Territórios associados (n)"], r["Ingredientes associados (n)"],
             r["Fonte"], oficial_ibge, r["Nome"], r["Descrição"]),
        )
        n_bio += 1

    # --- Livro/Fonte: 0 instâncias nesta versão (Dicionário v1.2, Seção 19 — esquema
    #     definido, população pendente para implementação futura). Nenhuma linha inserida.
    n_liv = 0

    # --- Relações (cabeçalho na linha 3) ---
    n_rel = 0
    for r in sheet_rows(wb["11. RELACOES"], header_row=3):
        conn.execute(
            """INSERT INTO relacoes (rel_id, origem_id, destino_id, tipo_relacao, fonte, pagina,
                confiabilidade, observacoes, data_criacao, peso, metodo_calculo_peso)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (r["REL_ID"], r["origem_id"], r["destino_id"], r["tipo_relacao"], r["fonte"], r["pagina"],
             r["confiabilidade"], r["observacoes"], r["data_criacao"],
             {"🟢 Confirmado em várias fontes": 0.95, "🟡 Confirmado em uma única fonte": 0.60,
              "🔵 Inferido": 0.30, "🔴 Pendente de validação": 0.10}.get(r["confiabilidade"], 0.30),
             f"Inicializado a partir do mapeamento determinístico de confiabilidade ({r['confiabilidade']}), Dicionário v1.2 Seção 20.2"),
        )
        n_rel += 1

    conn.commit()

    # --- Validação pós-carga contra o Relatório de Auditoria Sprint 2 ---
    esperado = {
        "ingrediente": 130, "receita": 136, "tecnica": 38, "povo": 17,
        "territorio": 18, "patrimonio": 35, "bioma": 7, "livro_fonte": 0,
    }
    obtido = {"ingrediente": n_ing, "receita": n_rec, "tecnica": n_tec, "povo": n_pov,
              "territorio": n_ter, "patrimonio": n_pat, "bioma": n_bio, "livro_fonte": n_liv}

    divergencias = [f"{k}: esperado {v}, obtido {obtido[k]}" for k, v in esperado.items() if obtido[k] != v]
    total_objetos = sum(obtido.values())
    if total_objetos != 381:
        divergencias.append(f"total_objetos: esperado 381, obtido {total_objetos}")
    if n_rel != 1585:
        divergencias.append(f"relacoes: esperado 1585, obtido {n_rel}")

    # Zero erros de FK
    fk_orfas = conn.execute(
        """SELECT COUNT(*) FROM relacoes
           WHERE origem_id NOT IN (SELECT id FROM objeto_universal)
              OR destino_id NOT IN (SELECT id FROM objeto_universal)"""
    ).fetchone()[0]
    if fk_orfas > 0:
        divergencias.append(f"FK inválida em relacoes: {fk_orfas} linha(s) apontam para ID inexistente")

    if divergencias:
        print("DIVERGÊNCIA DETECTADA — PARANDO CONFORME RESTRIÇÃO DA ORDEM 2:")
        for d in divergencias:
            print(" -", d)
        conn.close()
        sys.exit(1)

    print(f"ETL concluído sem divergências. Objetos: {total_objetos} (esperado 381). Relações: {n_rel} (esperado 1585).")
    print(f"Por catálogo: {obtido}")
    conn.close()


if __name__ == "__main__":
    main()
